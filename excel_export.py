"""
Excel and CSV export functionality.
"""

import pandas as pd
from typing import Dict, List
from io import BytesIO
from models import Scenario
from capex_engine import calculate_scenario_totals, aggregate_by_category, calculate_normalization_metrics
from formatting import format_cop, format_number, format_percentage


def export_to_excel(scenario: Scenario, totals: Dict[str, float]) -> BytesIO:
    """Export scenario to Excel with formulas."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Sheet 1: Inputs
    ws_inputs = wb.create_sheet("Inputs")
    ws_inputs.append(["Variable", "Valor"])
    ws_inputs.append(["P50 MWh/año", scenario.variables.p50_mwh_per_year])
    ws_inputs.append(["P90 MWh/año", scenario.variables.p90_mwh_per_year])
    ws_inputs.append(["Potencia AC (MW)", scenario.variables.ac_power_mw])
    ws_inputs.append(["Potencia DC (kWp)", scenario.variables.dc_power_mwp])
    ws_inputs.append(["Moneda", scenario.variables.currency])
    ws_inputs.append(["Tasa de cambio", scenario.variables.fx_rate])
    ws_inputs.append([])
    ws_inputs.append(["AIU Habilitado", scenario.aiu_config.enabled])
    ws_inputs.append(["AIU Admin %", scenario.aiu_config.admin_pct])
    ws_inputs.append(["AIU Imprev %", scenario.aiu_config.imprev_pct])
    ws_inputs.append(["AIU Util %", scenario.aiu_config.util_pct])
    ws_inputs.append([])
    ws_inputs.append(["IVA Recuperable", scenario.vat_config.vat_recoverable])
    ws_inputs.append(["IVA sobre Utilidad", scenario.vat_config.vat_on_utilidad_enabled])
    ws_inputs.append(["IVA Rate Utilidad %", scenario.vat_config.vat_rate_utilidad])
    
    # Sheet 2: Items
    ws_items = wb.create_sheet("Items")
    items_data = []
    for item in scenario.items:
        item_total = totals['item_totals'].get(item.item_id, {'base': 0.0, 'vat': 0.0, 'total': 0.0})
        scenario_kwp = scenario.variables.dc_power_mwp
        cop_per_kwp = item_total['total'] / scenario_kwp if scenario_kwp > 0 else 0.0
        
        items_data.append({
            'Código': item.item_code,
            'Ítem': item.name,
            'Categoría': item.category_code,
            'Cantidad': item.qty,
            'Unidad': item.unit,
            'Modo precio': 'Por unidad' if item.pricing_mode == 'UNIT' else 'Por kWp',
            'Precio': item.price,
            'IVA %': item.vat_rate,
            'Base': item_total['base'],
            'IVA': item_total['vat'],
            'Total': item_total['total'],
            'COP/kWp': cop_per_kwp,
            'Paga cliente': 'Sí' if item.client_pays else 'No'
        })
    
    df_items = pd.DataFrame(items_data)
    for r in dataframe_to_rows(df_items, index=False, header=True):
        ws_items.append(r)
    
    # Sheet 3: AIU Breakdown
    ws_aiu = wb.create_sheet("AIU Breakdown")
    ws_aiu.append(["Concepto", "Base", "Porcentaje", "Valor"])
    ws_aiu.append(["Base Administración", totals['aiu_base_a'], f"{scenario.aiu_config.admin_pct}%", totals['aiu_admin']])
    ws_aiu.append(["Base Imprevistos", totals['aiu_base_i'], f"{scenario.aiu_config.imprev_pct}%", totals['aiu_imprev']])
    ws_aiu.append(["Base Utilidad", totals['aiu_base_u'], f"{scenario.aiu_config.util_pct}%", totals['aiu_util']])
    ws_aiu.append([])
    ws_aiu.append(["Total AIU", "", "", totals['aiu_total']])
    if totals['vat_on_utilidad'] > 0:
        ws_aiu.append(["IVA sobre Utilidad", totals['aiu_util'], f"{scenario.vat_config.vat_rate_utilidad}%", totals['vat_on_utilidad']])
    
    # Sheet 4: Summary by Category
    ws_summary = wb.create_sheet("Summary by Category")
    cat_totals = aggregate_by_category(scenario, totals)
    summary_data = []
    for cat_code, cat_data in cat_totals.items():
        summary_data.append({
            'Categoría': cat_data['name'],
            'Base': cat_data['base'],
            'IVA': cat_data['vat'],
            'Total': cat_data['total']
        })
    
    df_summary = pd.DataFrame(summary_data)
    for r in dataframe_to_rows(df_summary, index=False, header=True):
        ws_summary.append(r)
    
    # Sheet 5: Totals
    ws_totals = wb.create_sheet("Totals")
    ws_totals.append(["Concepto", "Valor"])
    ws_totals.append(["Direct EPC Base", totals['direct_epc_base']])
    ws_totals.append(["Direct EPC IVA", totals['direct_epc_vat']])
    ws_totals.append(["Direct EPC Total", totals['direct_epc_total']])
    ws_totals.append([])
    ws_totals.append(["AIU Admin", totals['aiu_admin']])
    ws_totals.append(["AIU Imprev", totals['aiu_imprev']])
    ws_totals.append(["AIU Util", totals['aiu_util']])
    ws_totals.append(["Total AIU", totals['aiu_total']])
    if totals['vat_on_utilidad'] > 0:
        ws_totals.append(["IVA sobre Utilidad", totals['vat_on_utilidad']])
    ws_totals.append([])
    ws_totals.append(["EPC Base", totals['epc_base']])
    ws_totals.append(["EPC IVA", totals['epc_vat']])
    ws_totals.append(["EPC Total", totals['epc_total']])
    ws_totals.append([])
    ws_totals.append(["Client Base", totals['client_base']])
    ws_totals.append(["Client IVA", totals['client_vat']])
    ws_totals.append(["Client Total", totals['client_total']])
    ws_totals.append([])
    ws_totals.append(["Project Base", totals['project_base']])
    ws_totals.append(["Project IVA", totals['project_vat']])
    ws_totals.append(["Project Total", totals['project_total']])
    
    # Metrics
    metrics = calculate_normalization_metrics(totals['project_total'], scenario)
    ws_totals.append([])
    ws_totals.append(["COP/kWp", metrics.get('cop_per_kwp', 0)])
    ws_totals.append(["COP/MWac", metrics.get('cop_per_mwac', 0)])
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_items_to_csv(scenario: Scenario, totals: Dict[str, float]) -> str:
    """Export items to CSV."""
    items_data = []
    for item in scenario.items:
        item_total = totals['item_totals'].get(item.item_id, {'base': 0.0, 'vat': 0.0, 'total': 0.0})
        scenario_kwp = scenario.variables.dc_power_mwp
        cop_per_kwp = item_total['total'] / scenario_kwp if scenario_kwp > 0 else 0.0
        
        items_data.append({
            'Código': item.item_code,
            'Ítem': item.name,
            'Categoría': item.category_code,
            'Cantidad': item.qty,
            'Unidad': item.unit,
            'Precio': item.price,
            'IVA %': item.vat_rate,
            'Total': item_total['total'],
            'COP/kWp': cop_per_kwp,
            'Paga cliente': 'Sí' if item.client_pays else 'No'
        })
    
    df = pd.DataFrame(items_data)
    return df.to_csv(index=False, encoding='utf-8-sig')


def export_summary_to_csv(scenario: Scenario, totals: Dict[str, float]) -> str:
    """Export summary to CSV."""
    cat_totals = aggregate_by_category(scenario, totals)
    summary_data = []
    for cat_code, cat_data in cat_totals.items():
        summary_data.append({
            'Categoría': cat_data['name'],
            'Base': cat_data['base'],
            'IVA': cat_data['vat'],
            'Total': cat_data['total']
        })
    
    # Add totals
    summary_data.append({
        'Categoría': 'Total EPC',
        'Base': totals['epc_base'],
        'IVA': totals['epc_vat'],
        'Total': totals['epc_total']
    })
    summary_data.append({
        'Categoría': 'Total Proyecto',
        'Base': totals['project_base'],
        'IVA': totals['project_vat'],
        'Total': totals['project_total']
    })
    
    df = pd.DataFrame(summary_data)
    return df.to_csv(index=False, encoding='utf-8-sig')
